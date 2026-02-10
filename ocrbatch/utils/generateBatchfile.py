import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json

async def create_excel_file(data, path, length):
    """Create Excel file with specific formatting using openpyxl."""
    base_path = f'{os.getenv("ASSETS")}/{path}'
    try:
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        
        # Write data to worksheet
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Handle cell formatting
        merge_array = []
        row = 1
        
        # Handle string formatting and date formatting
        for row in range(1, ws.max_row + 1):
            # Format column C (date)
            c_cell = ws[f'C{row}']
            if c_cell.value and not isinstance(c_cell.value, datetime):
                # Handle date formatting if needed
                pass
            
            # Format column G (string)
            g_cell = ws[f'G{row}']
            if g_cell.value and g_cell.value != '':
                g_cell.number_format = '@'  # Force text format
        
        # Handle merging cells
        for row in range(1, ws.max_row + 1):
            a_cell = ws[f'A{row}']
            if not a_cell.value:
                # Merge cells in this row from column B to length
                ws.merge_cells(
                    start_row=row,
                    start_column=2,
                    end_row=row,
                    end_column=length
                )
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(base_path), exist_ok=True)
        
        # Save the workbook
        wb.save(base_path)
        
        logging.info(f"Excel file created successfully at path {base_path}")
        return True
        
    except Exception as e:
        logging.error(f"Error creating Excel file: {str(e)}")
        raise


async def generate_data(data, column):
    """
    Generate formatted data for Excel file.
    
    Args:
        data: List of data records
        column: Column configuration string
    """
    try:
        logging.info('Generate excel file data')
        bundle = []
        columns = json.loads(f"[{column}]")
        length = len(columns)
        
        # Add ID column at the beginning
        columns.insert(0, ['ID', 'nBundledetailid'])
        
        # Add headers
        bundle.append([col[0] for col in columns])
        
        # Add data rows
        for element in data:
            bundle.append([element.get(col[1], '') for col in columns])
            
        logging.info('File data Generating successful')
        return [bundle, length]
        
    except Exception as error:
        logging.error(f'Error while generating excel file data: {str(error)}')
        raise