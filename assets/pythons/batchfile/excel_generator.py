import os
import sys
import json
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple, Any
from datetime import datetime, timedelta

def get_date():
    
    # Get the current time with the desired timezone
    date = datetime.now()
    
    # Extract date components 
    year = date.year
    month = date.month
    day = date.day

    return f"{year:04d}-{month:02d}-{day:02d}"

    lgfile = get_date()

lgfile = get_date()
log_dir = f'{os.getenv('ROOT_PATH')}{lgfile}/'

log_file = os.path.join(log_dir, f'batchFile.log')

# Ensure the log directory exists
os.makedirs(log_dir, exist_ok=True)

class BatchProcessor:
    def __init__(self):
        self.log_app = 'BatchProcessor'
        self.logger = logging.getLogger(self.log_app)
        logging.basicConfig(filename=log_file, level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

    def generate_data(self, data: List[Dict], column_str: str) -> Tuple[List[List[Any]], int]:
        try:
            self.logger.info("Generating excel file data")
            bundle = []
            columns = json.loads(f"[{column_str}]")
            length = len(columns)
            
            # Add ID and nBundledetailid to columns
            columns.insert(0, ['ID', 'nBundledetailid'])
            
            # Add headers
            bundle.append([col[0] for col in columns])
            
            # Add data rows
            for element in data:
                bundle.append([element.get(col[1], '') for col in columns])
                
            self.logger.info("File data Generation successful")
            return bundle, length
            
        except Exception as error:
            self.logger.error(f"Error while generating excel file data: {str(error)}")
            raise error

    def create_excel_file(self, data: List[Dict], path: str, column_str: str) -> Dict:
        try:
            self.logger.info("Starting Excel file creation")
            bundle, length = self.generate_data(data, column_str)
            
            wb = Workbook()
            ws = wb.active
            
            # Add data to worksheet
            for row_data in bundle:
                ws.append(row_data)

            row = 1
            merge_array = []

            # Process cell types
            while True:
                cell_c = ws.cell(row=row, column=3)  # Column C
                cell_g = ws.cell(row=row, column=7)  # Column G
                
                if not cell_c.value and not cell_g.value:
                    break

                if cell_g.value and cell_g.value != '':
                    cell_g.data_type = 's'
                    
                row += 1

            # Process merges
            for row_idx in range(1, ws.max_row + 1):
                cell_a = ws.cell(row=row_idx, column=1)
                if not cell_a.value:
                    merge_array.append({
                        's': {'r': row_idx - 1, 'c': 1},
                        'e': {'r': row_idx - 1, 'c': length}
                    })

            # Apply merges
            for merge in merge_array:
                start_col = get_column_letter(merge['s']['c'] + 1)
                end_col = get_column_letter(merge['e']['c'] + 1)
                start_cell = f"{start_col}{merge['s']['r'] + 1}"
                end_cell = f"{end_col}{merge['e']['r'] + 1}"
                ws.merge_cells(f"{start_cell}:{end_cell}")

            # Create directory if needed
            directory = os.path.dirname(path)
            if not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)

            # Save workbook
            self.logger.info("Excel ready to save")
            wb.save(path)
            self.logger.info("File saved successfully")

            return {"msg": 1, "value": "Batch File in process success"}

        except Exception as error:
            self.logger.error(f"Error while creating excel: {str(error)}")
            return {"msg": -1, "value": "Batch File in process failed", "error": str(error)}

def main():
    if len(sys.argv) != 4:
        print("Usage: python batch_processor.py <data_json> <output_path> <column_string>")
        sys.exit(1)

    data_json = sys.argv[1]
    output_path = sys.argv[2]
    column_string = sys.argv[3]

    try:
        # Parse the JSON data
        data = json.loads(data_json)
        
        # Create processor and generate file
        processor = BatchProcessor()
        result = processor.create_excel_file(data, output_path, column_string)
        
        # Print result as JSON for NestJS to parse
        print(json.dumps(result))
        
    except Exception as e:
        print(json.dumps({
            "msg": -1,
            "value": "Batch File in process failed",
            "error": str(e)
        }))

if __name__ == "__main__":
    main()